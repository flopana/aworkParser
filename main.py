import calendar
import datetime
import json

from openpyxl import Workbook, load_workbook


def main():
    currentMonth = datetime.datetime.today().month
    currentYear = datetime.datetime.today().year
    try:
        currentMonth = int(input(f"Use the current month {currentMonth} or use a specific one? Press enter for default: {currentMonth}\n-> "))
    except ValueError:
        pass
    filename = input("Please enter the filename of the excel sheet inside the excel_sheets folder (including the file "
                     "ending!)\n-> ")
    wb = load_workbook(f"excel_sheets/{filename}")
    sheet = wb.sheetnames[0]
    sheet = wb[sheet]
    workedDict = {}

    for row in sheet.rows:
        if row[0].value == "User":
            continue

        date = row[1].value
        hoursWorked = round(row[4].value / 60 / 60, 2)

        if currentMonth == date.month and currentYear == date.year:
            try:
                workedDict[date] += hoursWorked
            except KeyError:
                workedDict |= {date: hoursWorked}

    exportWb = Workbook()
    sheet = exportWb.active
    sheet.title = f"Worktime Export {currentMonth}-{currentYear}"
    keys = list(workedDict.keys())
    keys.reverse()
    rowCounter = 2
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15

    sheet.cell(column=1, row=1, value="Datum")
    sheet.cell(column=2, row=1, value="Stunden")
    for key in keys:
        sheet.cell(column=1, row=rowCounter, value=key.strftime("%d.%m.%Y"))
        sheet.cell(column=2, row=rowCounter, value=round(workedDict[key], 2))
        print(key.strftime("%d.%m.%Y") + " | " + str(round(workedDict[key], 2)))
        rowCounter += 1

    workDaysForPrediction = json.loads(open("prediction.json").read())["workDays"]
    # If month for export is the actual month of executing this then add the prediction according to prediction.json
    if currentMonth == datetime.datetime.today().month:
        print("Adding prediction...")
        curr = datetime.datetime.today()
        daysLeft = calendar.monthrange(currentYear, currentMonth)[1] - keys[len(keys) - 1].day
        for i in range(daysLeft):
            curr = datetime.datetime(year=curr.year, month=curr.month, day=curr.day + 1)
            for tag in workDaysForPrediction:
                if curr.isoweekday() == tag["dayOfWeek"]:
                    sheet.cell(column=1, row=rowCounter, value=curr.strftime("%d.%m.%Y"))
                    sheet.cell(column=2, row=rowCounter, value=tag["hours"])
                    print(curr.strftime("%d.%m.%Y") + " | " + str(tag["hours"]))
                    rowCounter += 1

    sheet.cell(column=1, row=rowCounter + 1, value="Summe: ")
    sheet.cell(column=2, row=rowCounter + 1, value=f"=SUM(B2: B{rowCounter - 1})")

    exportFileName = f"exports/{currentMonth}_{currentYear}.xlsx"
    exportWb.save(exportFileName)
    print(f"\nExported to {exportFileName}")


if __name__ == '__main__':
    main()
